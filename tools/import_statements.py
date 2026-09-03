"""
Разовий імпорт банківських виписок у ZenMoney.

Читає виписки Monobank (CSV), A-Bank (PDF) і Raiffeisen (XLSX), зводить перекази
між власними рахунками в одну транзакцію, розставляє категорії за правилами
й формує payload для /v8/diff/.

За замовчуванням — dry-run: пише preview.csv, нікуди не відправляє.
Відправка тільки з явним --send.

payload.json — накопичувальний журнал відправок. Кожен --send дописує партію
з її id, тому відкотити можна будь-яку, а не тільки останню.

    python import_statements.py                 # прев'ю
    python import_statements.py --send          # запис у ZenMoney
    python import_statements.py --batches       # журнал відправок
    python import_statements.py --rollback-last # відкотити останню партію
    python import_statements.py --rollback      # відкотити все ще не відкочене

Суми й дати беруться з виписок як є — жодного LLM у цьому шляху.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import time
import uuid
from datetime import date, timedelta
from pathlib import Path

BASE = Path(__file__).parent.parent  # script lives in tools/, data dirs sit at repo root
CARDS = BASE / "Картки"
OWN = CARDS / "Особисті"
AD = CARDS / "Рекламні "
OUT = BASE / "import_out"

# --------------------------------------------------------------------------
# Мапінг виписка → рахунок у ZenMoney
# --------------------------------------------------------------------------

# Виписки згруповані за періодами, бо кожен імпортується окремо: червень-липень
# уже записаний (batch 1), і повторний прогін дав би дублі.
# Формат: (парсер, шлях, рахунок ZenMoney, контрольні суми витрат і зарахувань).
PERIODS: dict[str, list[tuple]] = {
    "jun-jul": [
        ("abank", OWN / "Абанк червень-липень.pdf", "abank green", -56669.01, 32658.71),
        ("mono", OWN / "Моно червень-липень.csv", "mono black", None, None),
        ("abank", AD / "Червень/АбанкЧервень.pdf", "abank yellow", -107385.22, 106026.98),
        ("abank", AD / "Липень/А-банк Липень.pdf", "abank yellow", -119801.66, 119801.66),
        ("raiff", AD / "Червень/Raif.xlsx", "yellow raiff", -106473.50, 104300.00),
        ("raiff", AD / "Липень/Райф липень.xlsx", "yellow raiff", -41517.86, 44460.62),
        ("mono", AD / "Червень/МоноЧервень.csv", "mono white", None, None),
        ("mono", AD / "Липень/Моно Липень.csv", "mono white", None, None),
    ],
    "aug": [
        ("abank", OWN / "abank green august.pdf", "abank green", -8026.78, 3481.46),
        ("mono", OWN / "mono black august.csv", "mono black", None, None),
        ("abank", AD / "abank yellow august.pdf", "abank yellow", -2500.00, 2500.00),
        ("raiff", AD / "raiff yellow.xlsx", "yellow raiff", -69626.19, 72936.14),
        ("mono", AD / "mono white august.csv", "mono white", None, None),
    ],
}
DEFAULT_PERIOD = "aug"
SOURCES = PERIODS[DEFAULT_PERIOD]

# Картка, що згадується в описі → рахунок-контрагент.
# Ключ — перші 4 + останні 4 цифри (маски в різних банках відрізняються:
# "4400 00** **** 1118", "44000*****731118", "414951******0442").
CARD_TO_ACCOUNT = {
    "44001118": "abank green",
    "41490442": "yellow raiff",
    "48746675": "mono white",
    "44413140": "mono white",
}

# --------------------------------------------------------------------------
# Бізнес-потоки Malvia.
#
# Рекламні витрати й виручка, що їх покриває, у ZenMoney не потрапляють —
# записується тільки різниця, тобто те, що реально лишилось на особисте.
# Наслідок: залишки на рахунках у ZenMoney НЕ збігатимуться з банківськими.
# Це свідомий вибір, а не помилка імпорту.
# --------------------------------------------------------------------------

NET_BUSINESS = True

BUSINESS_EXPENSE_RE = re.compile(
    r"facebk|facebook|keycrm|fopliashenko|ukr host|ukrhost|hosting"
    # Разові виплати за рекламу, які в описі виглядають як звичайні перекази.
    # Прив'язані до конкретних операцій 02.06.2026 на mono white — для нових
    # періодів такі треба звіряти окремо, за назвою вони не розпізнаються.
    r"|^олег б\.|^андрій м\.|^414943\*+3010",
    re.I,
)
BUSINESS_INCOME_RE = re.compile(r"novapay|ushan ihor", re.I)

# Готівка, внесена через термінали, — це наложка Malvia, а не особисті гроші.
CASH_IS_BUSINESS_REVENUE = True

# Рахунок і категорія, на які лягає порахована різниця.
NET_ACCOUNT = "mono white"
NET_TAG = "malvia"

# Рахунок, з якого йде готівка при внесенні через термінал.
CASH_ACCOUNT = "cash"

# Списання, що насправді є переказом на відомий рахунок (погашення розстрочки):
# (рахунок-джерело, регекс по опису) → рахунок-призначення.
APPLE_WATCH_ACCOUNT = "YA.UA Apple Watch Nika"

# Порядок важливий: перше співпадіння виграє. У серпні Монобанк перейменував
# «Платіж YA.UA» на «Щомісячний платіж YA.UA», і загальне правило почало
# перехоплювати його — тому конкретніше правило має стояти вище.
DIRECT_TRANSFERS: list[tuple[str, str, str]] = [
    ("mono black", r"ya\.ua", APPLE_WATCH_ACCOUNT),
    ("mono black", r"щомісячний платіж", "моно рассрочка на карту 07.01.26"),
]

# --------------------------------------------------------------------------
# Правила категоризації: (регекс по опису, MCC-коди, тег ZenMoney)
# Перше співпадіння виграє. Порядок має значення.
# --------------------------------------------------------------------------

# Підписки (Anthropic і перекази за них) — окремої категорії не заводимо,
# йдуть у наявну "платежи, комиссии".
SUBSCRIPTION_TAG = "платежи, комиссии"

RULES: list[tuple[str | None, set[str] | None, str]] = [
    (r"вероніка|вероника|zaitseva veronika", None, "никуся"),
    (r"київська школа економіки|\bkse\b", None, "доход"),      # стипендія KSE
    (r"виведення кешбеку|кешбек", None, "доход"),
    (r"яна роскош", None, "кафе и рестораны"),                 # повернули за піцу
    (r"anthropic", None, SUBSCRIPTION_TAG),
    (r"^артем [км]\.", None, SUBSCRIPTION_TAG),                # перекази за підписку
    (r"easypay", None, "вода"),
    (r"facebk|facebook", {"7311"}, "facebook ads"),
    (r"списання відсотків|за перевищення кредитного", None, "% кредит.лимит"),
    (r"гу дпс|гук в |податк|/\d{8}$", None, "налоги"),
    (r"multiplex|планета кіно|кінопалац", {"7832"}, "кино"),
    (r"київстар|kyivstar", None, "мобильный"),
    (r"knigarnya|книгарня|bookua", {"5942"}, "образование"),
    (r"vbet|favbet|parimatch|квазі-готівка", {"7995"}, "отдых и развлечения"),
    (r"аврора|duna|akvarel", {"5310", "5651", "5943"}, "покупки"),
    (r"gromadvbiralnia|жкг|водоканал", {"4900"}, "комуналка"),
    (r"liqpay\*ukr host|ukrhost|hosting", None, "интернет"),
    (r"mcdonald", None, "mcdonalds"),
    (r"bolt\s*food|glovo|raketa", None, "доставка еды"),
    (r"\bbolt\b|uklon|uber|\bтаксі\b", {"4121"}, "такси"),
    (r"метро|\bipay\b", {"4111"}, "метро"),
    (r"сільпо|silpo", None, "сильпо"),
    (r"\bатб\b|\batb\b", None, "атб"),
    (r"varus|варус", None, "varus"),
    (r"львівські круасани|lvivski", None, "львівські круасани"),
    (r"шаурм|shaurm|shauirm", None, "шаурма"),
    (r"велмарт|velmart", None, "велмарт"),
    (r"bikenow|скутер", {"7394"}, "скутери"),
    (r"barbershop|перукар", None, "забота о себе"),
    (r"\beva\b|watsons|аптек|pharm", {"5912", "5977"}, "здоровье и фитнес"),
    (r"portmone.*|киевстар|kyivstar|vodafone|lifecell", {"4814"}, "мобильный"),
    (r"інтернет|internet", {"4816"}, "интернет"),
    (r"novapay|нова пошта|nova poshta", {"4215"}, "malvia"),
    (r"wog|okko|shell|socar|бензин", {"5541", "5542"}, "бензин"),
    (r"jysk|epicentr|епіцентр|rozetka", {"5211", "5712"}, "покупки"),
    (r"apple|google|steam|netflix|spotify", {"5816", "5818"}, "покупки"),
    (None, {"4816"}, "интернет"),
    (None, {"5411", "5499", "5462"}, "продукты"),
    (None, {"5812", "5813", "5814", "5811"}, "кафе и рестораны"),
    (None, {"5912", "5977", "7230", "7298"}, "забота о себе"),
    (None, {"4111"}, "метро"),
    (None, {"4131"}, "транспорт"),   # 4131 — автобуси/міжміські, не метро
    (None, {"4121"}, "такси"),
]

# MCC 6010/6012 (фінустанови) навмисно НЕ мапиться: на вхідних переказах від
# людей цей код нічого не означає, а правило клеїло їм випадкову категорію
# замість того, щоб лишити їх за бортом. Реальні комісії ловить окреме
# правило "списання відсотків|за перевищення кредитного" вище.

FALLBACK_TAG = "без категории"

# Операції без правила все одно записуються — доходом або витратою з описом
# як є, у категорію FALLBACK_TAG. Постав True, щоб натомість викидати їх.
DROP_UNKNOWN = False

# Описи, що означають внесення готівки через термінал (переказ з cash).
# "Поповнення: DENYS IHOROVYCH KOROBCHY" — це теж внесення готівки через
# термінал. Формулювання відрізняється від переказів з власних карток
# ("MONO*KOROBCHYNSKYI DENYS", "ABNK*DENYS KOROBCHYNSKYI", "Від: Denys
# Korobchynskyi"), тому regex прицільний і їх не чіпає.
CASH_IN_RE = re.compile(
    r"термінал|usonmo|a2c visa cash|cash easypay|city24|denys ihorovych", re.I,
)

# Описи-кандидати на переказ між власними рахунками.
TRANSFER_RE = re.compile(
    r"своєї картки|білої картки|грошові перекази|поповнення[:\s]|переказ на картку|"
    r"mono\*|abnk\*|korobchynskyi|зарахування переказу|на свою картку|щомісячний платіж",
    re.I,
)
FIN_MCC = {"4829", "6010", "6012", "6011", "9999"}

# Формулювання вихідної / вхідної сторони переказу — мають бути узгоджені.
OUT_PHRASE_RE = re.compile(
    r"на свою картку|переказ на картку|грошові перекази|щомісячний платіж|"
    r"\d{4}[\s\*]", re.I,
)
IN_PHRASE_RE = re.compile(
    r"зі своєї картки|з білої картки|зарахування переказу|поповнення|від:|"
    r"mono\*|abnk\*", re.I,
)


# --------------------------------------------------------------------------
# Парсери виписок
# --------------------------------------------------------------------------

def _num_ua(s) -> float:
    return float(re.sub(r"[\s   ]", "", str(s)).replace(",", "."))


def parse_mono(path: Path, account: str) -> list[dict]:
    """
    Monobank CSV. Контрольних сум у файлі немає, тому перевіряємо інакше:
    залишок після кожної операції має дорівнювати залишку наступної (файл
    відсортований від свіжих до старих) плюс сума операції.
    """
    out = []
    with open(path, encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for i in range(len(rows) - 1):
        cur, nxt = rows[i], rows[i + 1]
        expected = (float(nxt["Залишок після операції"])
                    + float(cur["Сума в валюті картки (UAH)"]))
        if abs(expected - float(cur["Залишок після операції"])) > 0.011:
            raise SystemExit(
                f"{path.name}: розрив у залишку на {cur['Дата i час операції']} "
                f"({cur['Деталі операції']}): очікувано {expected:.2f}, "
                f"у файлі {cur['Залишок після операції']}. Виписка неповна."
            )
    for r in rows:
        dt = r["Дата i час операції"]
        out.append({
            "account": account,
            "date": f"{dt[6:10]}-{dt[3:5]}-{dt[0:2]}",
            "time": dt[11:],
            "desc": r["Деталі операції"].strip(),
            "mcc": r["MCC"].strip(),
            "uah": float(r["Сума в валюті картки (UAH)"]),
            "op_amount": float(r["Сума в валюті операції"]),
            "op_currency": r["Валюта"].strip(),
            "src": path.name,
        })
    return out


_DATE_RE = re.compile(r"^\s*(\d{2}\.\d{2}\.\d{4})\s*(.*)$")
_AMT_RE = re.compile(
    r"^\s*(?:-|\d{4}\*{4}\d{4})\s+(.*?)\s+(\d{4})\s+"
    r"(-?[\d\s   ]+,\d{2})\s+(-?[\d\s   ]+,\d{2})\s+([A-Z]{3})"
)


def parse_abank(path: Path, account: str) -> list[dict]:
    """A-Bank PDF. Опис операції переноситься на 2-3 рядки — склеюємо."""
    txt = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        capture_output=True, text=True, check=True,
    ).stdout
    lines = txt.splitlines()
    out, i = [], 0
    while i < len(lines):
        m = _DATE_RE.match(lines[i])
        if m and i + 1 < len(lines):
            a = _AMT_RE.match(lines[i + 1])
            if a:
                tail = ""
                if i + 2 < len(lines):
                    nxt = lines[i + 2]
                    if not _DATE_RE.match(nxt) and not _AMT_RE.match(nxt):
                        tail = nxt.strip()
                desc = " ".join(x for x in (m.group(2).strip(), a.group(1).strip(), tail) if x)
                tm = re.search(r"\b(\d{2}:\d{2})\b", desc)
                out.append({
                    "account": account,
                    "date": f"{m.group(1)[6:]}-{m.group(1)[3:5]}-{m.group(1)[:2]}",
                    "time": tm.group(1) if tm else "",
                    "desc": re.sub(r"\s*\d{2}:\d{2}\s*", " ", desc).strip(),
                    "mcc": a.group(2),
                    "uah": _num_ua(a.group(3)),
                    "op_amount": _num_ua(a.group(4)),
                    "op_currency": a.group(5),
                    "src": path.name,
                })
                i += 2
                continue
        i += 1
    return out


def parse_raiff(path: Path, account: str) -> list[dict]:
    import openpyxl

    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = next(
        (r for r in range(1, ws.max_row + 1)
         if str(ws.cell(r, 1).value or "").startswith("Дата і час")),
        None,
    )
    if hdr is None:
        raise ValueError(f"{path.name}: не знайдено рядок заголовка")
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 13)]
        if not v[0] or v[5] is None:
            continue
        ts = str(v[0])
        out.append({
            "account": account,
            "date": ts[:10],
            "time": ts[11:19],
            "desc": str(v[3]).strip(),
            "mcc": str(v[4] or "").strip(),
            "uah": float(v[5]),
            "op_amount": float(v[6]) if v[6] is not None else float(v[5]),
            "op_currency": str(v[7] or "UAH").strip(),
            "src": path.name,
        })
    return out


PARSERS = {"mono": parse_mono, "abank": parse_abank, "raiff": parse_raiff}


def extract_all(verbose: bool = True) -> list[dict]:
    txns = []
    for kind, path, account, exp_out, exp_in in SOURCES:
        rows = PARSERS[kind](path, account)
        got_out = sum(x["uah"] for x in rows if x["uah"] < 0)
        got_in = sum(x["uah"] for x in rows if x["uah"] > 0)
        # mono перевіряється безперервністю залишку всередині parse_mono
        status = "OK (залишок)" if kind == "mono" else "—"
        if exp_out is not None:
            ok = abs(got_out - exp_out) < 0.02 and abs(got_in - exp_in) < 0.02
            status = "OK" if ok else "РОЗБІЖНІСТЬ"
            if not ok:
                raise SystemExit(
                    f"{path.name}: суми не збігаються з випискою.\n"
                    f"  очікувано {exp_out:,.2f} / {exp_in:,.2f}\n"
                    f"  отримано  {got_out:,.2f} / {got_in:,.2f}"
                )
        if verbose:
            print(f"  {path.name[:34]:36} {account:13} {len(rows):5} шт  "
                  f"{got_out:>13,.2f} {got_in:>13,.2f}  {status}")
        txns += rows
    for i, x in enumerate(txns):
        x["i"] = i
    return txns


# --------------------------------------------------------------------------
# Крок 1: скасування
# --------------------------------------------------------------------------

# Префікси, що не є назвою мерчанта: банки додають їх по-різному в різних
# виписках, через них "Транспорт: LIQPAY*Avtolyuks" і "Повернення: LIQPAY*
# AVTOLYUKS" виглядають як різні контрагенти.
_PREFIX_RE = re.compile(
    r"^(скасування|повернення|транспорт|інші витрати|побутова техніка|"
    r"поповнення|перерахунок за)[.:\s]+", re.I,
)
CANCEL_RE = re.compile(r"^(скасування|повернення)\b", re.I)


def _merchant_key(desc: str) -> str:
    """Назва мерчанта без службових префіксів, згорнута у ключ для звірки."""
    s = _PREFIX_RE.sub("", desc).strip().lower()
    s = re.sub(r",.*$", "", s)
    return re.sub(r"[^a-zа-яіїєґ0-9]", "", s)[:8]


def apply_cancellations(txns: list[dict]) -> tuple[list[dict], list[str]]:
    """
    "Скасування. X" або "Повернення: X" гасить попереднє списання того самого
    мерчанта. Повна пара — обидва рядки геть. Часткова — оригінал зменшується
    на суму повернення, рядок повернення геть.
    """
    log = []
    drop: set[int] = set()
    cancels = [x for x in txns if CANCEL_RE.match(x["desc"]) and x["uah"] > 0]
    for c in cancels:
        key = _merchant_key(c["desc"])
        best = None
        for o in txns:
            if o["i"] in drop or o["uah"] >= 0 or o["account"] != c["account"]:
                continue
            if not key or _merchant_key(o["desc"]) != key:
                continue
            delta = (date.fromisoformat(c["date"]) - date.fromisoformat(o["date"])).days
            if not -1 <= delta <= 5:
                continue
            if abs(o["uah"]) < c["uah"] - 0.01:
                continue
            if best is None or o["date"] > best["date"]:
                best = o
        drop.add(c["i"])
        if best is None:
            log.append(f"  ! {c['date']} {c['account']:13} {c['desc'][:34]:36} "
                       f"+{c['uah']:,.2f} — оригінал не знайдено, рядок пропущено")
            continue
        if abs(abs(best["uah"]) - c["uah"]) < 0.01:
            drop.add(best["i"])
            log.append(f"  = {c['date']} {c['account']:13} {best['desc'][:34]:36} "
                       f"{c['uah']:,.2f} — повна пара, обидва рядки викинуто")
        else:
            was = best["uah"]
            best["uah"] += c["uah"]
            best["op_amount"] = best["uah"]
            log.append(f"  ~ {c['date']} {c['account']:13} {best['desc'][:34]:36} "
                       f"{was:,.2f} → {best['uah']:,.2f} (часткове повернення {c['uah']:,.2f})")
    return [x for x in txns if x["i"] not in drop], log


# --------------------------------------------------------------------------
# Крок 2: перекази між власними рахунками
# --------------------------------------------------------------------------

def _cards_in(desc: str) -> set[str]:
    out = set()
    for m in re.findall(r"\d{4}[\s\*]*[\d\*]{2,8}[\s\*]*\d{4}", desc):
        digits = re.sub(r"\D", "", m)
        if len(digits) >= 8:
            out.add(digits[:4] + digits[-4:])
    return out


def _hinted_account(x: dict) -> str | None:
    for card in _cards_in(x["desc"]):
        if card in CARD_TO_ACCOUNT:
            return CARD_TO_ACCOUNT[card]
    return None


def match_transfers(txns: list[dict]) -> tuple[list[dict], list[dict], list[str]]:
    """
    Зводить дві сторони одного переказу в одну транзакцію type=transfer.
    Пари шукає за сумою + датою (±3 дні), пріоритет — збіг номера картки в описі.
    """
    log = []
    cand = [x for x in txns
            if x["mcc"] in FIN_MCC or TRANSFER_RE.search(x["desc"])]
    cand = [x for x in cand if not CASH_IN_RE.search(x["desc"])]
    by_i = {x["i"]: x for x in cand}

    scored = []
    for a in cand:
        for b in cand:
            if a["i"] >= b["i"] or a["account"] == b["account"]:
                continue
            if abs(a["uah"] + b["uah"]) > 0.011:
                continue
            if (a["uah"] > 0) == (b["uah"] > 0):
                continue
            delta = abs((date.fromisoformat(a["date"]) - date.fromisoformat(b["date"])).days)
            if delta > 3:
                continue
            score = 0
            ha, hb = _hinted_account(a), _hinted_account(b)
            # згадка номера картки контрагента — найсильніший сигнал
            if ha == b["account"]:
                score += 10
            elif ha:
                score -= 8  # вказує на ІНШИЙ рахунок → майже напевно не пара
            if hb == a["account"]:
                score += 10
            elif hb:
                score -= 8
            # формулювання сторін узгоджені за напрямком
            src, dst = (a, b) if a["uah"] < 0 else (b, a)
            if OUT_PHRASE_RE.search(src["desc"]) and IN_PHRASE_RE.search(dst["desc"]):
                score += 6
            score -= delta
            scored.append((score, a["i"], b["i"]))

    scored.sort(key=lambda s: -s[0])
    used: set[int] = set()
    pairs = []
    for score, ia, ib in scored:
        if score < 0 or ia in used or ib in used:
            continue
        used.add(ia)
        used.add(ib)
        a, b = by_i[ia], by_i[ib]
        src, dst = (a, b) if a["uah"] < 0 else (b, a)
        pairs.append({
            "kind": "transfer", "date": src["date"], "amount": abs(src["uah"]),
            "from": src["account"], "to": dst["account"],
            "desc": f"{src['desc']} → {dst['desc']}",
            "confidence": "high" if score >= 9 else "low",
            "rows": [src["i"], dst["i"]],
        })
        if score < 9:
            log.append(f"  ? {src['date']} {abs(src['uah']):>10,.2f} "
                       f"{src['account']:13} → {dst['account']:13} "
                       f"score={score} | {src['desc'][:30]} | {dst['desc'][:30]}")

    rest = [x for x in txns if x["i"] not in used]
    return pairs, rest, log


def direct_transfers(txns: list[dict]) -> tuple[list[dict], list[dict]]:
    """Списання з відомим призначенням (погашення розстрочки тощо)."""
    out, rest = [], []
    for x in txns:
        hit = next(
            (dst for acc, pat, dst in DIRECT_TRANSFERS
             if x["account"] == acc and re.search(pat, x["desc"], re.I) and x["uah"] < 0),
            None,
        )
        if hit:
            out.append({
                "kind": "transfer", "date": x["date"], "amount": abs(x["uah"]),
                "from": x["account"], "to": hit,
                "desc": x["desc"], "confidence": "high", "rows": [x["i"]],
            })
        else:
            rest.append(x)
    return out, rest


def cash_transfers(txns: list[dict]) -> tuple[list[dict], list[dict]]:
    """Внесення готівки через термінал → переказ з рахунку cash."""
    out, rest = [], []
    for x in txns:
        if CASH_IN_RE.search(x["desc"]) and x["uah"] > 0:
            out.append({
                "kind": "transfer", "date": x["date"], "amount": x["uah"],
                "from": CASH_ACCOUNT, "to": x["account"],
                "desc": x["desc"], "confidence": "high", "rows": [x["i"]],
            })
        else:
            rest.append(x)
    return out, rest


# --------------------------------------------------------------------------
# Крок 3: категорії
# --------------------------------------------------------------------------

def net_business(
    txns: list[dict], cash: list[dict],
) -> tuple[list[dict], list[dict], list[str], list[dict]]:
    """
    Прибирає рекламні витрати й виручку, що їх фінансує, і замість них
    ставить одну транзакцію на місяць — різницю, яка лишилась на особисте.

    Якщо CASH_IS_BUSINESS_REVENUE, готівкові внески через термінали теж
    вважаються виручкою і перестають бути переказами з cash.
    """
    log = []
    biz_exp, biz_inc, rest = [], [], []
    for x in txns:
        # Знак не перевіряємо: повернення від Facebook («Повернення: FACEBK …»)
        # має зменшувати рекламні витрати, а не ставати особистим доходом.
        if BUSINESS_EXPENSE_RE.search(x["desc"]):
            biz_exp.append(x)
        elif x["uah"] > 0 and BUSINESS_INCOME_RE.search(x["desc"]):
            biz_inc.append(x)
        else:
            rest.append(x)

    if CASH_IS_BUSINESS_REVENUE:
        biz_inc += [{"date": c["date"], "uah": c["amount"], "desc": c["desc"],
                     "account": c["to"], "mcc": "", "i": -1} for c in cash]
        cash = []

    months = sorted({x["date"][:7] for x in biz_exp + biz_inc})
    netted = []
    for mo in months:
        inc = sum(x["uah"] for x in biz_inc if x["date"][:7] == mo)
        exp = sum(x["uah"] for x in biz_exp if x["date"][:7] == mo)
        net = inc + exp
        log.append(f"  {mo}: надійшло {inc:>12,.2f}  реклама {exp:>12,.2f}  "
                   f"→ чистими {net:>12,.2f}")
        if abs(net) < 0.01:
            continue
        last_day = (date.fromisoformat(f"{mo}-01") + timedelta(days=32)).replace(day=1) \
            - timedelta(days=1)
        netted.append({
            "kind": "income" if net > 0 else "expense",
            "date": last_day.isoformat(), "amount": abs(net),
            "account": NET_ACCOUNT, "desc": f"Malvia: чистими за {mo}",
            "mcc": "", "tag": NET_TAG, "confidence": "high", "rows": [],
        })
    log.append(f"  згорнуто {len(biz_exp)} рекламних витрат і {len(biz_inc)} надходжень "
               f"у {len(netted)} транзакц.")
    return rest, netted, log, cash


def pick_tag(x: dict) -> str:
    desc = x["desc"].lower()
    for pattern, mccs, tag in RULES:
        if pattern and re.search(pattern, desc):
            return tag
        if mccs and x["mcc"] in mccs and pattern is None:
            return tag
    return FALLBACK_TAG


def build_plain(txns: list[dict]) -> tuple[list[dict], list[dict]]:
    """Повертає (транзакції до запису, викинуті через відсутність правила)."""
    out, dropped = [], []
    for x in txns:
        # Правила застосовуються і до доходу: "Від: Вероніка Зайцева" має
        # лишитись у «Никусі». Те, під що правила немає, йде за борт.
        tag = pick_tag(x)
        unknown = tag == FALLBACK_TAG
        item = {
            "kind": "expense" if x["uah"] < 0 else "income",
            "date": x["date"], "amount": abs(x["uah"]),
            "account": x["account"], "desc": x["desc"], "mcc": x["mcc"],
            "tag": tag, "confidence": "low" if unknown else "high",
            "rows": [x["i"]],
        }
        if unknown and DROP_UNKNOWN:
            dropped.append(item)
        else:
            out.append(item)
    return out, dropped


# --------------------------------------------------------------------------
# Крок 4: payload для ZenMoney
# --------------------------------------------------------------------------

def to_zenmoney(items: list[dict], state: dict) -> list[dict]:
    accounts = state["accounts"]
    tags = state["tags"]
    user_id = state["userId"]
    now = int(time.time())
    missing_acc, missing_tag = set(), set()
    payload = []

    for it in items:
        def acc_id(name):
            # state.json індексує рахунки за title.lower() — звіряємось так само,
            # щоб назва з великих літер ("YA.UA Apple Watch Nika") теж знайшлась.
            a = accounts.get(name.lower().strip())
            if a is None:
                missing_acc.add(name)
                return None, 0
            return a["id"], a["instrument"]

        if it["kind"] == "transfer":
            src_id, src_cur = acc_id(it["from"])
            dst_id, dst_cur = acc_id(it["to"])
            income, outcome = it["amount"], it["amount"]
            inc_acc, out_acc = dst_id, src_id
            inc_cur, out_cur = dst_cur, src_cur
            tag_ids = []
        else:
            a_id, cur = acc_id(it["account"])
            inc_acc = out_acc = a_id
            inc_cur = out_cur = cur
            if it["kind"] == "income":
                income, outcome = it["amount"], 0.0
            else:
                income, outcome = 0.0, it["amount"]
            t = tags.get(it["tag"])
            if t is None:
                missing_tag.add(it["tag"])
                tag_ids = []
            else:
                tag_ids = [t["id"] if isinstance(t, dict) else t]

        payload.append({
            "id": str(uuid.uuid4()), "changed": now, "created": now,
            "user": user_id, "deleted": False,
            "incomeInstrument": inc_cur, "incomeAccount": inc_acc, "income": income,
            "outcomeInstrument": out_cur, "outcomeAccount": out_acc, "outcome": outcome,
            "tag": tag_ids, "merchant": None, "payee": None, "originalPayee": None,
            "comment": it["desc"][:200], "date": it["date"],
            "mcc": int(it["mcc"]) if it.get("mcc", "").isdigit() else None,
            "reminderMarker": None, "opIncome": None, "opIncomeInstrument": None,
            "opOutcome": None, "opOutcomeInstrument": None,
            "latitude": None, "longitude": None,
            "incomeBankID": None, "outcomeBankID": None, "qrCode": None,
        })

    if missing_acc:
        raise SystemExit(f"Немає таких рахунків у ZenMoney: {sorted(missing_acc)}\n"
                         f"Доступні: {sorted(accounts)}")
    if missing_tag:
        print(f"  ⚠️  немає категорій, підуть без тега: {sorted(missing_tag)}")
    return payload


# --------------------------------------------------------------------------

PAYLOAD = OUT / "payload.json"


def load_ledger() -> list[dict]:
    """
    payload.json — журнал усіх відправок, а не знімок останньої.
    Кожен елемент: {batch, sent_at, count, rolled_back, transactions}.

    Старий формат (плаский список транзакцій) підхоплюється як batch 1,
    щоб не втратити можливість відкотити те, що вже записано.
    """
    if not PAYLOAD.exists():
        return []
    data = json.load(open(PAYLOAD))
    if data and isinstance(data[0], dict) and "transactions" not in data[0]:
        print("   payload.json у старому форматі — переношу в журнал як batch 1")
        ledger = [{
            "batch": 1, "sent_at": None, "sent_at_iso": "невідомо (до журналу)",
            "count": len(data), "status": "sent", "rolled_back": None,
            "transactions": data,
        }]
        save_ledger(ledger)  # закріплюємо міграцію, щоб вона не повторювалась
        return ledger
    return data


def save_ledger(ledger: list[dict]) -> None:
    json.dump(ledger, open(PAYLOAD, "w"), ensure_ascii=False, indent=1)


def append_batch(payload: list[dict]) -> list[dict]:
    ledger = load_ledger()
    now = int(time.time())
    ledger.append({
        "batch": max((b["batch"] for b in ledger), default=0) + 1,
        "sent_at": now,
        "sent_at_iso": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(now)),
        "count": len(payload),
        # Журнал пишеться ДО відправки: якщо звʼязок обірветься на півдорозі,
        # id для відкату вже збережені. Статус переходить у "sent" після того,
        # як API підтвердив запис.
        "status": "pending",
        "rolled_back": None,
        "transactions": payload,
    })
    save_ledger(ledger)
    return ledger


def confirm_batch(ledger: list[dict]) -> None:
    ledger[-1]["status"] = "sent"
    save_ledger(ledger)


def rollback(args) -> None:
    """
    Відкат: ті самі транзакції відправляються ще раз з deleted=True.
    Форма об'єкта та сама, що й при записі, тож ZenMoney просто позначає
    їх видаленими.

    За замовчуванням відкочуються всі партії, які ще не відкочені.
    З --rollback-last — тільки найсвіжіша.
    """
    import asyncio

    sys.path.insert(0, str(BASE))
    from zenmoney import ZenMoneyClient  # noqa: E402

    ledger = load_ledger()
    if not ledger:
        raise SystemExit(f"Немає {PAYLOAD} — нема чого відкочувати.")

    live = [b for b in ledger if not b["rolled_back"]]
    if not live:
        raise SystemExit("Усі партії вже відкочені — нема чого робити.")

    targets = [live[-1]] if args.rollback_last else live
    print("Журнал відправок:")
    for b in ledger:
        mark = "відкочено" if b["rolled_back"] else \
               ("→ ВІДКОЧУ" if b in targets else "лишається")
        print(f"  batch {b['batch']}  {b['sent_at_iso']}  {b['count']:>4} шт   {mark}")

    txns = [t for b in targets for t in b["transactions"]]
    print(f"\nВидаляю {len(txns)} транзакцій із {len(targets)} партій.")

    now = int(time.time())
    for t in txns:
        t["deleted"] = True
        t["changed"] = now
    asyncio.run(ZenMoneyClient().add_transactions(txns))

    for b in targets:
        b["rolled_back"] = now
        b["rolled_back_iso"] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(now))
    save_ledger(ledger)
    print("Готово. Партії позначені відкоченими в журналі.")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--send", action="store_true",
                    help="реально записати в ZenMoney (без прапорця — тільки прев'ю)")
    ap.add_argument("--cash-is-revenue", action="store_true",
                    help="вважати готівку з терміналів виручкою Malvia, а не переказом з cash")
    ap.add_argument("--refresh", action="store_true",
                    help="оновити state.json з ZenMoney перед записом (потрібна мережа)")
    ap.add_argument("--limit", type=int, metavar="N",
                    help="відправити тільки перші N транзакцій — для пробного запуску")
    ap.add_argument("--rollback", action="store_true",
                    help="видалити з ZenMoney усі ще не відкочені партії "
                         "(за id з журналу import_out/payload.json)")
    ap.add_argument("--rollback-last", action="store_true",
                    help="відкотити тільки останню партію")
    ap.add_argument("--batches", action="store_true",
                    help="показати журнал відправок і вийти")
    ap.add_argument("--period", choices=sorted(PERIODS), default=DEFAULT_PERIOD,
                    help=f"який період імпортувати (за замовчуванням {DEFAULT_PERIOD})")
    ap.add_argument("--only", metavar="REGEX",
                    help="взяти лише транзакції, чий опис підходить під регекс — "
                         "щоб дописати окремі операції в уже імпортований період")
    args = ap.parse_args()

    globals()["SOURCES"] = PERIODS[args.period]

    if args.batches:
        ledger = load_ledger()
        if not ledger:
            print("Журнал порожній — нічого не відправлялось.")
            return
        print(f"{'batch':>5}  {'коли':19}  {'шт':>5}  стан")
        for b in ledger:
            if b["rolled_back"]:
                state = f"відкочено {b.get('rolled_back_iso', '')}"
            elif b.get("status") == "pending":
                state = "⚠️  відправку не підтверджено — перевір у ZenMoney"
            else:
                state = "живе в ZenMoney"
            print(f"{b['batch']:>5}  {b['sent_at_iso']:19}  {b['count']:>5}  {state}")
        live = sum(b["count"] for b in ledger if not b["rolled_back"])
        print(f"\nживих транзакцій: {live}")
        return

    if args.rollback or args.rollback_last:
        rollback(args)
        return

    if args.cash_is_revenue:
        globals()["CASH_IS_BUSINESS_REVENUE"] = True

    OUT.mkdir(exist_ok=True)

    print(f"1. Читаю виписки за період: {args.period}")
    txns = extract_all()
    print(f"   разом {len(txns)} рядків\n")

    print("2. Скасування")
    txns, log = apply_cancellations(txns)
    for line in log:
        print(line)
    print(f"   лишилось {len(txns)} рядків\n")

    print("3. Перекази між власними рахунками")
    pairs, txns, tlog = match_transfers(txns)
    cash, txns = cash_transfers(txns)
    direct, txns = direct_transfers(txns)
    pairs += direct
    for line in tlog:
        print(line)
    print(f"   спарено {len(pairs)} переказів (з них {len(direct)} за прямим правилом), "
          f"готівкових внесків {len(cash)}")
    print(f"   лишилось {len(txns)} звичайних операцій\n")

    netted = []
    if NET_BUSINESS:
        print("4. Бізнес-потоки Malvia (згортаю в різницю)")
        txns, netted, nlog, cash = net_business(txns, cash)
        for line in nlog:
            print(line)
        print()

    print("5. Категорії")
    plain, dropped = build_plain(txns)
    items = pairs + cash + plain + netted
    items.sort(key=lambda x: x["date"])

    if args.only:
        only_re = re.compile(args.only, re.I)
        before = len(items)
        items = [i for i in items if only_re.search(i["desc"])]
        print(f"   фільтр --only {args.only!r}: {len(items)} із {before}")
        for i in items:
            print(f"     {i['date']}  {i['amount']:>9,.2f}  "
                  f"{i.get('account') or i.get('from')}  {i['desc'][:44]}")
    if dropped:
        agg: dict[str, list] = {}
        for d in dropped:
            k = d["desc"][:34].strip()
            agg.setdefault(k, [0, 0.0, d["kind"]])
            agg[k][0] += 1
            agg[k][1] += d["amount"]
        print(f"   за бортом (немає правила): {len(dropped)} шт на "
              f"{sum(d['amount'] for d in dropped):,.2f} ₴")
        for k, (n, s, kind) in sorted(agg.items(), key=lambda x: -x[1][1])[:15]:
            print(f"     {n:3}x {kind:7} {k:36} {s:>11,.2f}")
        with open(OUT / "dropped.csv", "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["дата", "тип", "сума", "рахунок", "опис", "mcc"])
            for d in sorted(dropped, key=lambda x: x["date"]):
                w.writerow([d["date"], d["kind"], f"{d['amount']:.2f}",
                            d["account"], d["desc"][:80], d["mcc"]])
    print(f"   до запису: {len(items)}\n")

    preview = OUT / "preview.csv"
    with open(preview, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["дата", "тип", "сума", "рахунок", "куди", "категорія",
                    "опис", "mcc", "перевірити"])
        for it in items:
            w.writerow([
                it["date"], it["kind"], f"{it['amount']:.2f}",
                it.get("account") or it.get("from", ""), it.get("to", ""),
                it.get("tag", ""), it["desc"][:80], it.get("mcc", ""),
                "ТАК" if it.get("confidence") == "low" else "",
            ])
    json.dump(items, open(OUT / "items.json", "w"), ensure_ascii=False, indent=1)
    print(f"   прев'ю: {preview}")

    exp = sum(i["amount"] for i in items if i["kind"] == "expense")
    inc = sum(i["amount"] for i in items if i["kind"] == "income")
    trf = sum(i["amount"] for i in items if i["kind"] == "transfer")
    print(f"\n   витрати   {exp:>13,.2f} ₴")
    print(f"   доходи    {inc:>13,.2f} ₴")
    print(f"   перекази  {trf:>13,.2f} ₴")
    print(f"   транзакцій до запису: {len(items)}")

    if not args.send:
        print("\nDry-run. Перевір preview.csv, далі запускай з --send.")
        return

    sys.path.insert(0, str(BASE))
    from zenmoney import ZenMoneyClient, load_state  # noqa: E402
    import asyncio

    client = ZenMoneyClient()
    if args.refresh:
        print("\n   оновлюю state.json з ZenMoney …")
        state = asyncio.run(client.fetch_initial_data())
        print(f"   рахунків {len(state['accounts'])}, категорій {len(state['tags'])}")
    else:
        state = load_state()
    if not state.get("accounts"):
        raise SystemExit("state.json порожній — запусти з --refresh.")
    payload = to_zenmoney(items, state)
    if args.limit:
        payload = payload[:args.limit]
        print(f"   пробний запуск: тільки перші {len(payload)} транзакцій")
    ledger = append_batch(payload)
    print(f"   записано в журнал як batch {ledger[-1]['batch']} "
          f"(усього партій {len(ledger)})")

    print(f"\n6. Відправляю {len(payload)} транзакцій у ZenMoney")
    asyncio.run(client.add_transactions(payload))
    confirm_batch(ledger)
    print("   готово")


if __name__ == "__main__":
    main()
